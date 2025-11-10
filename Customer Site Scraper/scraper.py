from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import pandas as pd
import openpyxl
from dotenv import load_dotenv
import os

########## FUNCTIONS START ##########
def pah_cat_food_scrape(url, row_dict, driver):
    driver.get(url)
    wait = WebDriverWait(driver, 5)
    row_dict = row_dict
    
    try:
        price_whole = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "purchase-type-selector_price__kb9w9"))).text
        price = price_whole.replace('£', '')
        price = pd.to_numeric(price)
        product_name = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "preview_title-base-product__RDtS0"))).text
        
        row_dict['product_name'].value = product_name
        row_dict['last_run'].value = pd.Timestamp.now().strftime('%Y/%m/%d')
        old_price = row_dict['price'].value
        
        try:
            rating = wait.until(EC.presence_of_element_located((By.XPATH, '/html/body/div[1]/main/section/div/div[1]/div/div/div/div/div/p'))).text
            rating = rating.split(' ')[0]  # Extract the numeric part of the rating
            rating = round(pd.to_numeric(rating), 1)  # Convert to float and round to 1 decimal place
            row_dict['rating'].value = rating
        except:
            row_dict['rating'].value = 'No rating found' 
        
        if old_price == '' or old_price is None:
            row_dict['price'].value = price
            row_dict['status_since_last_run'].value = 'Price updated from empty or None'
        elif price < old_price:
            row_dict['price'].value = price
            row_dict['status_since_last_run'].value = f'Price dropped by {old_price - price}'
        elif price > old_price:
            row_dict['price'].value = price
            row_dict['status_since_last_run'].value = f'Price increased by {price - old_price}'
        else:
            row_dict['status_since_last_run'].value = 'No price change'
    except:
        row_dict['status_since_last_run'].value = 'Error: Potential broken URL'
        row_dict['price'].value = ''
        row_dict['last_run'].value = ''
        row_dict['product_name'].value = 'ERROR: Potential broken URL'
        row_dict['rating'].value = ''
    
    return row_dict
########## FUNCTIONS END ##########
load_dotenv()
file_path = os.getenv("scraped_data_workbook_path")
safety_file_path = os.getenv("scraped_data_workbook_safety_path")
wb = openpyxl.load_workbook(file_path)
options = webdriver.ChromeOptions()
options.add_argument('--headless')
options.add_argument('--disable-images')
driver = webdriver.Chrome(options=options)

for sheet in wb.sheetnames:
    headers = ['url', 'product_name', 'price', 'rating', 'last_run', 'status_since_last_run']
    sheet = wb[sheet]
    if sheet.title.lower() == 'pets at home cat food':
        for row in sheet.iter_rows(min_row=2, values_only=False):
            row_dict = dict(zip(headers, row))
            url = row_dict['url'].value
            row_dict = pah_cat_food_scrape(url, row_dict, driver)
    else:
        print(f"Skipping sheet: {sheet.title} as it is not 'Pets at Home Cat Food'")
    print(f"Processed sheet: {sheet.title}")
    
# Save the workbook after updating
wb.save(file_path)
wb.close()
driver.quit()

prompt_status = True
while prompt_status:
    prompt = input('Would you like to create a copy of the workbook for safety? (yes/no): ').lower()

    if prompt == 'yes':
        # Create a copy of the workbook and save it with a new name
        wb.save(safety_file_path)
        prompt_status = False
    elif prompt == 'no':
        print("No copy created. Proceeding without a backup.")
        prompt_status = False
    else:
        print("Invalid input. Please enter 'yes' or 'no'.")